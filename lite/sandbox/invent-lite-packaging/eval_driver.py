"""
Eval driver for the lite packaging spikes (Artifact A = one-file exe,
Artifact B = portable embedded-Python folder).

Runs with the system Python 3.14 (only for driving HTTP + process control —
the artifacts themselves run in their own runtimes with a sanitized PATH).

Cases:
  1  zero-Python launch  — sanitized PATH, PYTHONHOME/PYTHONPATH unset,
                           poll /health <=30s, record cold-start seconds
  2  big real permit w/ Thai path — upload, pageinfo/1, page/1, page/20,
                           thumb/45, export-xlsx round-trip via openpyxl
  3  double-launch       — 2nd instance binds next free port; 1st keeps serving

Usage:  python eval_driver.py A|B
Writes: results_<label>.json in this directory.
"""
import io
import json
import os
import re
import subprocess
import sys
import threading
import time
import urllib.request
from pathlib import Path

SB = Path(__file__).resolve().parent
SANITIZED_PATH = r"C:\Windows\system32;C:\Windows;C:\Windows\System32\Wbem;C:\Windows\System32\WindowsPowerShell\v1.0"
PDF = SB / "แบบทดสอบ" / "โครงการทดสอบ_แบบก่อสร้าง.pdf"

URL_RE = re.compile(r"http://127\.0\.0\.1:(\d+)/")


def artifact_cmd(label):
    if label == "A":
        return [str(SB / "dist" / "BMA-Plan-Lite-A.exe")]
    return ["cmd", "/c", str(SB / "portable" / "run.bat")]


def sanitized_env():
    env = dict(os.environ)
    env["PATH"] = SANITIZED_PATH
    env.pop("PYTHONHOME", None)
    env.pop("PYTHONPATH", None)
    env["PYTHONUNBUFFERED"] = "1"
    env["BMA_LITE_NO_BROWSER"] = "1"  # honored by wrapper A only; B has no flag
    return env


class Instance:
    def __init__(self, label):
        self.proc = subprocess.Popen(
            artifact_cmd(label), env=sanitized_env(),
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            cwd=str(SB), text=True, encoding="utf-8", errors="replace")
        self.port = None
        self.lines = []
        self.t0 = time.perf_counter()
        self._reader = threading.Thread(target=self._read, daemon=True)
        self._reader.start()

    def _read(self):
        for line in self.proc.stdout:
            self.lines.append(line.rstrip())
            m = URL_RE.search(line)
            if m and self.port is None:
                self.port = int(m.group(1))

    def wait_port(self, timeout=15):
        t = time.time()
        while self.port is None and time.time() - t < timeout:
            time.sleep(0.1)
        return self.port

    def wait_health(self, timeout=30):
        """Poll /health until 200. Returns seconds from process start, or None."""
        deadline = time.time() + timeout
        while time.time() < deadline:
            if self.port is not None:
                try:
                    with urllib.request.urlopen(
                            f"http://127.0.0.1:{self.port}/health", timeout=2) as r:
                        if r.status == 200:
                            json.loads(r.read())
                            return round(time.perf_counter() - self.t0, 2)
                except Exception:
                    pass
            time.sleep(0.25)
        return None

    def kill(self):
        subprocess.run(["taskkill", "/T", "/F", "/PID", str(self.proc.pid)],
                       capture_output=True)
        try:
            self.proc.wait(timeout=10)
        except Exception:
            pass


def http(method, url, data=None, headers=None, timeout=120):
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    t0 = time.perf_counter()
    with urllib.request.urlopen(req, timeout=timeout) as r:
        body = r.read()
        return r.status, body, round(time.perf_counter() - t0, 2)


def multipart_upload(url, filepath):
    boundary = "----bmaLiteSpike"
    fname = Path(filepath).name  # Thai filename — goes into the part header UTF-8
    body = io.BytesIO()
    body.write(f"--{boundary}\r\n".encode())
    body.write(("Content-Disposition: form-data; name=\"file\"; "
                f"filename=\"{fname}\"\r\n").encode("utf-8"))
    body.write(b"Content-Type: application/pdf\r\n\r\n")
    body.write(Path(filepath).read_bytes())
    body.write(f"\r\n--{boundary}--\r\n".encode())
    return http("POST", url, data=body.getvalue(),
                headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})


def working_sets():
    """tasklist snapshot: {pid: (name, mem_kb)} for python/BMA processes."""
    out = subprocess.run(["tasklist", "/FO", "CSV", "/NH"], capture_output=True,
                         text=True, encoding="cp874", errors="replace").stdout
    res = {}
    for line in out.splitlines():
        parts = [p.strip('"') for p in line.split('","')]
        if len(parts) >= 5 and ("python" in parts[0].lower() or "bma" in parts[0].lower()):
            res[parts[1]] = (parts[0], parts[4])
    return res


def run_case2(port, results):
    base = f"http://127.0.0.1:{port}"
    c2 = {"steps": {}}
    try:
        st, body, dt = multipart_upload(f"{base}/upload", PDF)
        up = json.loads(body)
        c2["steps"]["upload"] = {"status": st, "sec": dt, "pages": up.get("pages"),
                                 "name_echo": up.get("name")}
        cid = up["case_id"]
        st, _, dt = http("GET", f"{base}/pageinfo/1?case_id={cid}")
        c2["steps"]["pageinfo_1"] = {"status": st, "sec": dt}
        st, b1, dt1 = http("GET", f"{base}/page/1?case_id={cid}")
        c2["steps"]["page_1"] = {"status": st, "sec": dt1, "bytes": len(b1)}
        st, b20, dt = http("GET", f"{base}/page/20?case_id={cid}")
        c2["steps"]["page_20"] = {"status": st, "sec": dt, "bytes": len(b20)}
        st, bt, dt = http("GET", f"{base}/thumb/45?case_id={cid}")
        c2["steps"]["thumb_45"] = {"status": st, "sec": dt, "bytes": len(bt)}
        payload = json.dumps({
            "rows": [{"page": 1, "category": "GFA", "semanticTag": "base_area",
                      "kind": "poly", "area": 123.45, "count": 1}],
            "summary": [{"category": "GFA", "total": 123.45}],
        }).encode()
        st, xb, dt = http("POST", f"{base}/export-xlsx", data=payload,
                          headers={"Content-Type": "application/json"})
        xlsx_ok = False
        sheets = None
        if st == 200:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(xb))
            sheets = wb.sheetnames
            xlsx_ok = wb["Measurements"].max_row == 2 and "Summary" in sheets
        c2["steps"]["export_xlsx"] = {"status": st, "sec": dt, "bytes": len(xb),
                                      "openpyxl_parses": xlsx_ok, "sheets": sheets}
        c2["upload_plus_first_render_sec"] = round(
            c2["steps"]["upload"]["sec"] + dt1, 2)
        c2["pass"] = (all(s.get("status") == 200 for s in c2["steps"].values())
                      and xlsx_ok and up.get("pages") == 45)
        c2["working_sets_after"] = working_sets()
    except Exception as e:
        c2["error"] = repr(e)
        c2["pass"] = False
    results["case2"] = c2


def main():
    label = sys.argv[1]
    results = {"artifact": label, "ts": time.strftime("%Y-%m-%d %H:%M:%S")}

    # ---- Case 1: zero-Python launch -------------------------------------
    inst = Instance(label)
    port = inst.wait_port()
    cold = inst.wait_health(30) if port else None
    results["case1"] = {
        "pass": cold is not None,
        "port": port,
        "cold_start_sec": cold,
        "sanitized_path": SANITIZED_PATH,
        "stdout_head": inst.lines[:6],
    }
    if cold is None:
        results["case1"]["stdout_all"] = inst.lines[-40:]
        inst.kill()
        print(json.dumps(results, ensure_ascii=False, indent=1))
        (SB / f"results_{label}.json").write_text(
            json.dumps(results, ensure_ascii=False, indent=1), encoding="utf-8")
        return

    # ---- Case 2: big permit, Thai name (reuse running instance 1) --------
    run_case2(port, results)

    # ---- Case 3: double launch ------------------------------------------
    inst2 = Instance(label)
    port2 = inst2.wait_port()
    cold2 = inst2.wait_health(30) if port2 else None
    # instance 1 must still serve during instance 2's life
    try:
        st1, _, _ = http("GET", f"http://127.0.0.1:{port}/health", timeout=5)
    except Exception as e:
        st1 = repr(e)
    results["case3"] = {
        "pass": (cold2 is not None and port2 != port and st1 == 200),
        "port_instance1": port, "port_instance2": port2,
        "instance2_cold_start_sec": cold2,
        "instance1_health_during_instance2": st1,
        "instance2_stdout_head": inst2.lines[:6],
    }
    inst2.kill()
    inst.kill()
    (SB / f"results_{label}.json").write_text(
        json.dumps(results, ensure_ascii=False, indent=1), encoding="utf-8")
    print(json.dumps(results, ensure_ascii=False, indent=1))


if __name__ == "__main__":
    main()
