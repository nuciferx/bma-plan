#!/usr/bin/env python3
"""
BMA-Plan Lite — real export-endpoint tests (S1/S5 hardening, 2026-07-02 audit).

Fills the test hole: no prior test POSTs a real payload to /export-xlsx or
/export-pdf-overlay and validates the produced bytes (test_export_submenu.py
only stubs client-side dlPost). Boots server_lite via a uvicorn thread and
drives it over plain HTTP with `requests` (server-side test → no Playwright).

Cases:
  T1 xlsx happy    — small real payload -> 200, PK zip magic, openpyxl-openable
  T2 overlay happy — upload 1-page PDF -> case_id -> poly+label -> 200, %PDF, 1 page
  T3 oversized     — pages>MAX_EXPORT_PAGES / pts>MAX_PTS / objs>MAX_OBJECTS -> 400
  T4 malformed     — unknown case_id -> 400; non-numeric page key -> 400 (was a
                     latent 500: sorted(key=int) crashed before the fix); 1e12 coord -> 400
  T5 xlsx rows     — rows>MAX_XLSX_ROWS -> 400

Emits LITE_EXPORT_ENDPOINTS_OK on success; exits 1 on any failure.

    py -3 lite/tests/test_export_endpoints.py
"""
import io
import socket
import sys
import threading
import time
from pathlib import Path

LITE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(LITE))

import fitz          # PyMuPDF
import requests
import uvicorn

from server_lite import (app as lite_app, MAX_EXPORT_PAGES, MAX_OBJECTS_PER_PAGE,
                         MAX_PTS_PER_OBJECT, MAX_XLSX_ROWS)

try:
    import openpyxl
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False


def _free_port(start=8590):
    for p in range(start, start + 60):
        with socket.socket() as s:
            if s.connect_ex(("127.0.0.1", p)) != 0:
                return p
    raise RuntimeError("no free port")


failures = []


def check(label, cond, detail=""):
    if cond:
        print(f"  PASS {label}")
    else:
        msg = f"FAIL {label}" + (f": {detail}" if detail else "")
        failures.append(msg)
        print(f"  {msg}")


def _make_pdf_bytes(pages=1, w=300, h=400):
    doc = fitz.open()
    for i in range(pages):
        pg = doc.new_page(width=w, height=h)
        pg.insert_text(fitz.Point(50, 100), f"P{i+1}", fontsize=24)
    b = doc.tobytes()
    doc.close()
    return b


def t1_xlsx_happy(base):
    payload = {
        "rows": [
            {"page": 1, "category": "GFA", "semanticTag": "gross_floor_area",
             "kind": "area", "area": 2000.0, "count": None},
            {"page": 1, "category": "Deduction", "semanticTag": "void",
             "kind": "area", "area": 100.0, "count": None},
            {"page": 2, "category": "Parking", "semanticTag": "parking",
             "kind": "count", "area": None, "count": 1},
        ],
        "summary": [{"category": "GFA", "total": 1900.0}],
    }
    r = requests.post(base + "/export-xlsx", json=payload, timeout=30)
    check("T1 status==200", r.status_code == 200, f"got {r.status_code}")
    check("T1 xlsx content-type",
          "spreadsheetml.sheet" in r.headers.get("content-type", ""),
          r.headers.get("content-type"))
    check("T1 PK zip magic", r.content[:2] == b"PK", r.content[:4])
    if HAVE_OPENPYXL:
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        check("T1 has Measurements sheet", "Measurements" in wb.sheetnames, wb.sheetnames)
        ws = wb["Measurements"]
        check("T1 row count==4", ws.max_row == 4, f"got {ws.max_row}")  # header + 3 data
    else:
        print("  (openpyxl unavailable — magic + content-type only)")


def t2_overlay_happy(base):
    up = requests.post(base + "/upload",
                       files={"file": ("t.pdf", _make_pdf_bytes(1), "application/pdf")},
                       timeout=30)
    check("T2 upload 200", up.status_code == 200, f"got {up.status_code}")
    cid = up.json().get("case_id")
    check("T2 got case_id", bool(cid), up.text)
    payload = {"case_id": cid, "pages": {"1": {
        "objects": [{"kind": "poly", "counting": False,
                     "pts": [{"x": 10, "y": 10}, {"x": 60, "y": 10},
                             {"x": 60, "y": 60}, {"x": 10, "y": 60}],
                     "color": "#ff3b30", "label": "GFA 2000.00 m2"}],
        "annotations": [{"type": "ann_text", "pt": {"x": 20, "y": 20},
                         "text": "note", "color": "#ff453a"}]}}}
    r = requests.post(base + "/export-pdf-overlay", json=payload, timeout=30)
    check("T2 status==200", r.status_code == 200, f"got {r.status_code}")
    check("T2 %PDF magic", r.content[:4] == b"%PDF", r.content[:8])
    if r.content[:4] == b"%PDF":
        out = fitz.open(stream=r.content, filetype="pdf")
        check("T2 page count==1", len(out) == 1, f"got {len(out)}")
        out.close()
    return cid


def t3_oversized(base, cid):
    big_pages = {str(i): {"objects": [], "annotations": []}
                 for i in range(1, MAX_EXPORT_PAGES + 2)}
    r = requests.post(base + "/export-pdf-overlay",
                      json={"case_id": cid, "pages": big_pages}, timeout=60)
    check("T3a pages>MAX -> 400", r.status_code == 400, f"got {r.status_code}")

    many_pts = [{"x": 1.0, "y": 1.0}] * (MAX_PTS_PER_OBJECT + 1)
    r = requests.post(base + "/export-pdf-overlay",
                      json={"case_id": cid, "pages": {"1": {
                          "objects": [{"kind": "poly", "pts": many_pts, "color": "#f00"}],
                          "annotations": []}}}, timeout=30)
    check("T3b pts>MAX -> 400", r.status_code == 400, f"got {r.status_code}")

    many_objs = [{"kind": "poly", "pts": [{"x": 0, "y": 0}, {"x": 1, "y": 1}],
                  "color": "#f00"}] * (MAX_OBJECTS_PER_PAGE + 1)
    r = requests.post(base + "/export-pdf-overlay",
                      json={"case_id": cid, "pages": {"1": {
                          "objects": many_objs, "annotations": []}}}, timeout=30)
    check("T3c objs>MAX -> 400", r.status_code == 400, f"got {r.status_code}")


def t4_malformed(base, cid):
    r = requests.post(base + "/export-pdf-overlay",
                      json={"case_id": "does-not-exist",
                            "pages": {"1": {"objects": [], "annotations": []}}}, timeout=30)
    check("T4 unknown case_id -> 400", r.status_code == 400, f"got {r.status_code}")

    r = requests.post(base + "/export-pdf-overlay",
                      json={"case_id": cid,
                            "pages": {"abc": {"objects": [], "annotations": []}}}, timeout=30)
    check("T4 non-numeric page key -> 400", r.status_code == 400, f"got {r.status_code}")

    r = requests.post(base + "/export-pdf-overlay",
                      json={"case_id": cid, "pages": {"1": {
                          "objects": [{"kind": "poly",
                                       "pts": [{"x": 1e12, "y": 0}, {"x": 1, "y": 1},
                                               {"x": 2, "y": 2}], "color": "#f00"}],
                          "annotations": []}}}, timeout=30)
    check("T4 1e12 coord -> 400", r.status_code == 400, f"got {r.status_code}")


def t5_xlsx_rows(base):
    rows = [{"page": 1, "category": "x", "semanticTag": "t", "kind": "area",
             "area": 1.0, "count": None}] * (MAX_XLSX_ROWS + 1)
    r = requests.post(base + "/export-xlsx", json={"rows": rows, "summary": []}, timeout=60)
    check("T5 rows>MAX -> 400", r.status_code == 400, f"got {r.status_code}")


def main():
    port = _free_port()
    cfg = uvicorn.Config(lite_app, host="127.0.0.1", port=port, log_level="error")
    server = uvicorn.Server(cfg)
    threading.Thread(target=server.run, daemon=True).start()
    time.sleep(2.0)
    base = f"http://127.0.0.1:{port}"
    try:
        print("T1 xlsx happy path...")
        t1_xlsx_happy(base)
        print("T2 overlay happy path...")
        cid = t2_overlay_happy(base)
        print("T3 oversized rejected...")
        t3_oversized(base, cid)
        print("T4 malformed rejected...")
        t4_malformed(base, cid)
        print("T5 xlsx rows cap...")
        t5_xlsx_rows(base)
    finally:
        server.should_exit = True
        time.sleep(0.4)

    if failures:
        print()
        for f in failures:
            print("FAIL:", f)
        print("LITE_EXPORT_ENDPOINTS_FAIL")
        sys.exit(1)
    print()
    print("LITE_EXPORT_ENDPOINTS_OK")


if __name__ == "__main__":
    main()
